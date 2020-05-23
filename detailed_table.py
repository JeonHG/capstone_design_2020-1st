import os
import math
from xfoil import XFoil
from xfoil.model import Airfoil
import numpy as np
import pandas as pd

# load .dat files
array6409 = np.loadtxt("NACA6409.dat", skiprows=1)
array2412 = np.loadtxt("NACA2412.dat", skiprows=1)
naca6409 = Airfoil(x=array6409[:, 0], y=array6409[:, 1])
naca2412 = Airfoil(x=array2412[:, 0], y=array2412[:, 1])

# xf = XFoil()
# xf.airfoil = naca6409
# xf.Re = 1e6
# xf.max_it er = 40
# print(xf.a(10)[0])

# load blade profile - fixed values : pitch, chord_length
base_dir = r'C:\Users\USER\dev\mechanics'
csv_file = 'blade_profile.csv'
csv_dir = os.path.join(base_dir, csv_file)
df = pd.read_csv(csv_dir)

# df to dict
dfdict = df.to_dict(orient="index")
density = 1.225
# df_collection = []

def total_df(angular_velocity):
    w = angular_velocity
    for key, value in dfdict.items():
        value["blade_velocity"] = value['r_position'] * w
        value["relative_velocity"] = round(math.sqrt(value["blade_velocity"]**2 + value["wind_velocity"]**2),2)
        value["arctan"] = math.degrees(math.atan2(value["wind_velocity"], value["blade_velocity"]))
        aoa = round(value["arctan"] - value["pitch_angle"], 2)
        value["angle_of_attack"] = aoa
        re_n = round(value["relative_velocity"] * value["chord_length"] / 0.00001511, 3)
        value["Reynolds_number"] = re_n
        xf = XFoil()
        if key < 13: 
            xf.airfoil = naca6409
        else:
            xf.airfoil = naca2412
        xf.Re = re_n
        xf.max_iter = 100
        value["Cl"], value["Cd"], value["Cm"], value["Cp"] = xf.a(aoa)
        force_reference = 0.5 * density * value["relative_velocity"]**2
        if math.isnan(value["Cl"]):
            value["torque"] = 0
        else:
            lift = value["Cl"] * force_reference * 0.0125 * value['chord_length']
            drag = value["Cd"] * force_reference * 0.0125 * value['chord_length']
            value["lift"] = lift
            value["drag"] = drag
            value["torque"] = value["r_position"] * lift * math.sin(math.radians(value["pitch_angle"])) 
            # value["torque"] = value["r_position"] * (lift * math.sin(math.radians(value["pitch_angle"])) - drag * math.cos(math.radians(value["pitch_angle"])))
        
    detailed_df = pd.DataFrame.from_dict(dfdict, orient="index")
    print(detailed_df)
    return detailed_df

from openpyxl import Workbook, load_workbook

for w in [50, 60, 70, 80, 90]:

    df = total_df(w)
    sh_name = f"angular_{w}"
    
    try:
        book = load_workbook("detailed-table.xlsx")
        writer = pd.ExcelWriter('detailed-table.xlsx', engine="openpyxl", mode='a')
        writer.book = book
            # df = pd.DataFrame.from_dict(total_dict, orient='index', columns=['torque(N*m)'])
        writer.sheets = {ws.title: ws for ws in book.worksheets}
            # if sh_name in writer.sheets:
        if sh_name in writer.sheets:
            df.to_excel(writer, sheet_name=sh_name, startcol=0, startrow=writer.sheets[sh_name].max_row, index=True, header=True, index_label=f'{w}rad/s')
        else:
            df.to_excel(writer, sheet_name=sh_name, startrow=0, index=True, header=True, index_label=f'{w}rad/s')
        writer.save()
        writer.close()
    except FileNotFoundError:
        wb = Workbook()
        wb.save("detailed-table.xlsx")
        wb.close()
        book = load_workbook("detailed-table.xlsx")
        writer = pd.ExcelWriter('detailed-table.xlsx', engine='openpyxl', mode='w')
        writer.book = book
        # df = pd.DataFrame.from_dict(total_dict, orient='index', columns=['torque(N*m)'])
        df.to_excel(writer, sheet_name=f'angular_{w}', startrow=0, startcol=0, index=True, header=True, index_label=f"{w}rad/s")
        writer.save()
        writer.close()