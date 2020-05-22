import os
import math
from xfoil import XFoil
from xfoil.model import Airfoil
import numpy as np
import pandas as pd

# load .dat files
array6409 = np.loadtxt("NACA6409.dat", skiprows=1)
array2412 = np.loadtxt("NACA2412.dat", skiprows=1)
naca6409 = Airfoil(x=array6409[:, 0], y=array6409[:, 1])
naca2412 = Airfoil(x=array2412[:, 0], y=array2412[:, 1])

# xf = XFoil()
# xf.airfoil = naca6409
# xf.Re = 1e6
# xf.max_it er = 40
# print(xf.a(10)[0])

# load blade profile - fixed values : pitch, chord_length
base_dir = r'C:\Users\USER\dev\mechanics'
csv_file = 'blade_profile.csv'
csv_dir = os.path.join(base_dir, csv_file)
df = pd.read_csv(csv_dir)

# df to dict
dfdict = df.to_dict(orient="index")
density = 1.225
df_collection = []


def get_torque(angular_velocity):
    torque_sum_small = 0
    torque_sum_large = 0
    w = angular_velocity
    for key, value in dfdict.items():
        value["blade_velocity"] = value['r_position'] * w
        value["relative_velocity"] = round(math.sqrt(value["blade_velocity"]**2 + value["wind_velocity"]**2),2)
        value["arctan"] = math.degrees(math.atan2(value["wind_velocity"], value["blade_velocity"]))
        aoa = round(value["arctan"] - value["pitch_angle"], 2)
        value["angle_of_attack"] = aoa
        re_n = round(value["relative_velocity"] * value["chord_length"] / 0.00001511, 2)
        value["Reynolds_number"] = re_n
        xf = XFoil()
        if key < 13: 
            xf.airfoil = naca6409
        else:
            xf.airfoil = naca2412
        xf.Re = re_n
        xf.max_iter = 400
        value["Cl"], value["Cd"], value["Cm"], value["Cp"] = xf.a(aoa)
        force_reference = 0.5 * density * value["relative_velocity"]**2
        if math.isnan(value["Cl"]):
            value["torque"] = 0
        else:
            lift = value["Cl"] * force_reference * 0.0125 * value['chord_length']
            drag = value["Cd"] * force_reference * 0.0125 * value['chord_length']
            value["torque"] = value["r_position"] * (lift * math.sin(math.radians(value["arctan"])) - drag * math.cos(math.radians(value["arctan"])))
        if key < 13:
            torque_sum_small += value["torque"]
        else:
            pass
        if key > 0:
            torque_sum_large += value["torque"]
        else:
            pass
    df2 = pd.DataFrame.from_dict(dfdict, orient="index")
    df_collection.append(df2)
    torque_sum_avg = 0.5 * (torque_sum_small + torque_sum_large)
    return torque_sum_avg


torque_dict = {}
for w in range(0,200,10):
    torque_dict[w] = get_torque(w)
torque_df = pd.DataFrame.from_dict(torque_dict, orient='index', columns=["Torque (N*m)"])
torque_df.index.name = 'rad/s'
print(torque_df)

from openpyxl import Workbook, load_workbook

try:
    book = load_workbook("performance.xlsx")
    existing = pd.read_excel("performance.xlsx", sheet_name="torque-w")
    
    writer = pd.ExcelWriter('performance.xlsx', engine="openpyxl", mode='a')
    writer.book = book
    df = pd.DataFrame.from_dict(torque_dict, orient='index', columns=['torque(N*m)'])
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    for sheetname in writer.sheets:
        df.to_excel(writer, sheet_name=sheetname, startrow=0, startcol=writer.sheets[sheetname].max_column, index=True, header=True, index_label='rad/s')
    writer.save()
    writer.close()
except FileNotFoundError:
    wb = Workbook()
    wb.save("./performance.xlsx")
    wb.close()
    book = load_workbook("performance.xlsx")
    writer = pd.ExcelWriter('performance.xlsx', engine='openpyxl', mode='w')
    writer.book = book
    df = pd.DataFrame.from_dict(torque_dict, orient='index', columns=['torque(N*m)'])
    df.to_excel(writer, sheet_name='torque-w', startrow=0, startcol=0, index=True, header=True, index_label="rad/s")
    writer.save()
    writer.close()