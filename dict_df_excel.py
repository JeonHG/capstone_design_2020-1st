from openpyxl import Workbook, load_workbook
import pandas as pd


def dict_to_xlsx(dictionary, prefix, word):
    df = pd.DataFrame.from_dict(dictionary, orient="index")

    if type(word) == str:
        sh_name = word
        label = "rad/s"
    else:
        sh_name = f"{word}rad_per_s"
        label = "m"
    prefix = prefix + "-"
    try:
        book = load_workbook(f"{prefix}detailed-table.xlsx")
        writer = pd.ExcelWriter(f'{prefix}detailed-table.xlsx', engine="openpyxl", mode='a')
        writer.book = book
            # df = pd.DataFrame.from_dict(total_dict, orient='index', columns=['torque(N*m)'])
        writer.sheets = {ws.title: ws for ws in book.worksheets}
            # if sh_name in writer.sheets:
        if sh_name in writer.sheets:
            df.to_excel(writer, sheet_name=sh_name, startcol=0, startrow=writer.sheets[sh_name].max_row, index=True, header=True, index_label=label)
        else:
            df.to_excel(writer, sheet_name=sh_name, startrow=0, index=True, header=True, index_label=label)
        writer.save()
        writer.close()
    except FileNotFoundError:
        wb = Workbook()
        wb.save(f"{prefix}detailed-table.xlsx")
        wb.close()
        book = load_workbook(f"{prefix}detailed-table.xlsx")
        writer = pd.ExcelWriter(f'{prefix}detailed-table.xlsx', engine='openpyxl', mode='w')
        writer.book = book
        # df = pd.DataFrame.from_dict(total_dict, orient='index', columns=['torque(N*m)'])
        df.to_excel(writer, sheet_name=sh_name, startrow=0, startcol=0, index=True, header=True, index_label=label)
        writer.save()
        writer.close()
